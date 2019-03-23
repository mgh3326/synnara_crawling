import json

import pandas as pd
from openpyxl import load_workbook
import os
import platform
from random import *
from time import sleep

from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement

from dvd.model.DvdRes import DvdRes
import xlsxwriter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# for webdriver chromedriver 2.35 linux64
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('no-sandbox')
chrome_options.add_argument('headless')
# chrome_options.add_argument('window-size=1100x400')
chrome_options.add_argument('disable-gpu')
if platform.system() == "Windows":
    try:
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver.exe'), chrome_options=chrome_options)
    except ConnectionResetError:
        sleep(float("{0:.2f}".format(uniform(1, 2))))  # Time in seconds.
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver.exe'), chrome_options=chrome_options)
else:
    try:
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver'), chrome_options=chrome_options)
    except ConnectionResetError:
        sleep(float("{0:.2f}".format(uniform(1, 2))))  # Time in seconds.
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver'), chrome_options=chrome_options)
driver.implicitly_wait(3)


# url에 접근한다.

def Parse(_url):
    driver.get(_url)
    dvd = DvdRes()
    ##productImage
    productImage: WebElement = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[1]/p/img""")
    dvd.productImage = productImage.get_attribute('src')
    ##detailPageImage
    detailPageImage = driver.find_element_by_xpath("""//*[@id="tab_detail2"]/div[2]""")
    images = detailPageImage.find_elements_by_tag_name('img')
    for image in images:
        dvd.detailPageImage.append(image.get_attribute('src'))

    title = driver.find_element_by_xpath("""//*[@id="container"]/div[3]/div/ul[2]/li[1]""")
    dvd.title = title.text
    originalPrice = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[1]/table/tbody/tr[1]/td/span[1]""")
    dvd.originalPrice = originalPrice.text

    discountPrice = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[1]/table/tbody/tr[1]/td/span[2]""")
    dvd.discountPrice = discountPrice.text
    sold = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[2]""")
    images = sold.find_elements_by_tag_name('img')
    dvd.sold = images[0].get_attribute('alt')
    castMember = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[1]/td""")
    dvd.castMember = castMember.text
    director = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[2]/td[1]""")
    dvd.director = director.text

    screenRatio = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[3]/td""")
    dvd.screenRatio = screenRatio.text
    sound = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[4]/td""")
    dvd.sound = sound.text
    sound = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[4]/td""")

    dvd.sound = sound.text
    subtitle = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[6]/td""")
    dvd.subtitle = subtitle.text
    productionDistribution = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[2]/td[2]""")
    dvd.productionDistribution = productionDistribution.text
    regionCode = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[7]/td[1]""")
    dvd.regionCode = regionCode.text
    rating = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[7]/td[2]""")
    dvd.rating = rating.text
    runningTime = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[8]/td[1]""")
    dvd.runningTime = runningTime.text
    media = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[8]/td[2]""")
    dvd.media = media.text

    details = driver.find_element_by_xpath(
        """//*[@id="PROD_DESCR_DIV"]""")
    dvd.details = details.text

    return dvd


def ExcelWrite(_excel_path):
    if not os.path.exists(_excel_path):
        # Workbook is created

        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(_excel_path)
        worksheet = workbook.add_worksheet()

        worksheet.write(0, 0, '제품 이미지')
        worksheet.write(0, 1, '상세페이지 이미지')
        worksheet.write(0, 2, '제목')
        worksheet.write(0, 3, '원래 가격')
        worksheet.write(0, 4, '할인 가격')
        worksheet.write(0, 5, '품절')
        worksheet.write(0, 6, '출연자')
        worksheet.write(0, 7, '감독')
        worksheet.write(0, 8, '화면비율')
        worksheet.write(0, 9, '음향')
        worksheet.write(0, 10, '더빙')
        worksheet.write(0, 11, '자막')
        worksheet.write(0, 12, '제작배급')
        worksheet.write(0, 13, '지역코드')
        worksheet.write(0, 14, '등급')
        worksheet.write(0, 15, '런닝타임')
        worksheet.write(0, 16, '미디어')
        worksheet.write(0, 17, '상세정보')

        workbook.close()

    book = load_workbook(_excel_path)
    writer = pd.ExcelWriter(_excel_path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    # df = pd.DataFrame(json.dumps(dvd.__dict__))
    if len(dvd.detailPageImage) == 0:
        dvd.detailPageImage.append(" ")

    myjson = json.dumps(dvd.__dict__)

    print(myjson)

    df = pd.read_json((json.dumps(dvd.__dict__)))
    # df = pd.DataFrame.from_dict((json.dumps(dvd.__dict__)))
    for sheetname in writer.sheets:
        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                    header=False,
                    encoding="cp949")

    writer.save()


url = "http://www.synnara.co.kr/sp/sp121Main.do?categoryId=CT21004401&productId=P000421135#.XJXpYJgzZPY"

dvd = Parse(url)
print(dvd.title)
driver.quit()
excel_path = "./out3.xlsx"
ExcelWrite(excel_path)
