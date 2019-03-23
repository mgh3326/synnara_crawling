import json

import pandas as pd
from openpyxl import load_workbook
import os
import platform
from random import *
from time import sleep

from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement

import xlsxwriter

from record.model.RecordRes import RecordRes

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# for webdriver chromedriver 2.35 linux64
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('no-sandbox')
chrome_options.add_argument('headless')
# chrome_options.add_argument('window-size=1100x400')
chrome_options.add_argument('disable-gpu')
if platform.system() == "Windows":
    try:
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver.exe'), chrome_options=chrome_options)
    except ConnectionResetError:
        sleep(float("{0:.2f}".format(uniform(1, 2))))  # Time in seconds.
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver.exe'), chrome_options=chrome_options)
else:
    try:
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver'), chrome_options=chrome_options)
    except ConnectionResetError:
        sleep(float("{0:.2f}".format(uniform(1, 2))))  # Time in seconds.
        driver = webdriver.Chrome(os.path.join(
            BASE_DIR, '../chromedriver/chromedriver'), chrome_options=chrome_options)
driver.implicitly_wait(3)


# url에 접근한다.

def RecordParse(_url):
    driver.get(_url)
    record = RecordRes()
    ##productImage
    productImage: WebElement = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[1]/p[1]/img""")
    record.productImage = productImage.get_attribute('src')
    ##detailPageImage
    detailPageImage = driver.find_element_by_xpath("""//*[@id="tab_detail2"]/div[2]""")
    images = detailPageImage.find_elements_by_tag_name('img')
    image_list = []
    for image in images:
        image_list.append(image.get_attribute('src'))
    record.detailPageImage.append(','.join(map(str, image_list)))
    title = driver.find_element_by_xpath("""//*[@id="container"]/div[3]/div/ul[2]/li[1]""")
    record.title = title.text
    originalPrice = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[1]/table/tbody/tr[1]/td/span[1]""")
    record.originalPrice = originalPrice.text

    discountPrice = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[1]/table/tbody/tr[1]/td/span[2]""")
    record.discountPrice = discountPrice.text
    sold = driver.find_element_by_xpath(
        """//*[@id="container"]/div[3]/div/div[1]/div[2]/div[2]""")
    images = sold.find_elements_by_tag_name('img')
    record.sold = images[0].get_attribute('alt')
    artist = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[1]/td""")
    record.artist = artist.text
    publisher = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[2]/td[2]""")
    record.publisher = publisher.text
    label = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[3]/td[2]/a""")
    record.label = label.text
    productCode = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div[1]/table/tbody/tr[4]/td[1]""")
    if productCode.text.find("/") > -1:
        record.productCode = productCode.text.split("/")[0].strip()
        record.barcode = productCode.text.split("/")[1].strip()
    else:
        record.productCode = productCode.text
    releaseDate = driver.find_element_by_xpath(
        """//*[@id="tab_detail1"]/div/table/tbody/tr[4]/td[2]""")
    if releaseDate.text.find('[') > -1:
        record.releaseDate = releaseDate.text.split('[')[0].strip()
        record.weight = releaseDate.text.split('[')[1].split(']')[0].strip()  # [이거 있으면 뒤에 ] 있겠지 에러 안 뜨게 해야 되나
    else:
        record.releaseDate = releaseDate.text
    details = driver.find_element_by_xpath(
        """//*[@id="PROD_DESCR_DIV"]""")
    record.details = details.text

    return record


def ExcelWrite(_excel_path):
    if not os.path.exists(_excel_path):
        # Workbook is created

        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook(_excel_path)
        worksheet = workbook.add_worksheet()

        worksheet.write(0, 0, '제품 이미지')
        worksheet.write(0, 1, '상세페이지 이미지')
        worksheet.write(0, 2, '제목')
        worksheet.write(0, 3, '원래 가격')
        worksheet.write(0, 4, '할인 가격')
        worksheet.write(0, 5, '품절')
        worksheet.write(0, 6, '아티스트')
        worksheet.write(0, 7, '제작사')
        worksheet.write(0, 8, '레이블')
        worksheet.write(0, 9, '제품코드')
        worksheet.write(0, 10, '바코드')
        worksheet.write(0, 11, '발매일')
        worksheet.write(0, 12, '무게')
        worksheet.write(0, 13, '상세정보')

        workbook.close()

    book = load_workbook(_excel_path)
    writer = pd.ExcelWriter(_excel_path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    # df = pd.DataFrame(json.dumps(dvd.__dict__))
    if len(record.detailPageImage) == 0:
        record.detailPageImage.append(" ")

    myjson = json.dumps(record.__dict__)

    print(myjson)
    # df = pd.DataFrame({'count': json.dumps(record.__dict__)})
    jsonData = json.loads(myjson)
    df = pd.DataFrame(jsonData)

    # df = pd.read_json(jsonData)
    # df = pd.DataFrame.from_dict((json.dumps(dvd.__dict__)))
    for sheetname in writer.sheets:
        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                    header=False,
                    encoding="cp949")

    writer.save()


url = "http://www.synnara.co.kr/sp/sp120Main.do?categoryId=CT22101001&productId=P000422727#.XJX-dJgzZPa"

record = RecordParse(url)
print(record.productCode)
driver.quit()
excel_path = "./out5.xlsx"
ExcelWrite(excel_path)
